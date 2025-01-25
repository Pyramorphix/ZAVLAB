import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

# Maximum row count in the spreadsheet. Feel free to change if needed
MAX_ROW_COUNT = 100

# This is for one very unethical hack in the code for calculating lsd error. (see: format_error())
# Defines a maximum absolute value of the order of numbers in user input fields.
# If equals to x, the program will correctly work with numbers from 10^{-x} to 10^{x+1} - 1.
# I think, no one is crazy enough to type in numbers like 0.0000000005, so 9 is more than enough.
# Feel free to change, but keep in mind that lsd excel formulas' complexity will scale proportionally 
MAX_EXPONENT_ABS = 9

# The row at which data begins in a spreadsheet.
# The first row is for experiment titles, and the second one is for field labels.
# Therefore, we start from the third.
# Feel free to increase, decreasing will most likely break something
DATA_BEGINNING_ROW = 3


"""
Here we make the basement for out spreadsheet
"""
# =====================================================================================================================

class Spreadsheet:
    """
    The basement for creating a spreadsheet. Basically a list consisting of fields.

    Each field is either for user input (type='gathered') or for calculated values (type='calculated').
    Every field has the label and measure unit.

    User input fields additionally have the 'error' element for entering the experimental error.
    It can be either a certain value (e.g. '0.005') or a formula depending on measured value (e.g. '2% + 0.04')

    Calculated values have the 'formula' element for the calculation formula.
    It should be typed in using excel-readable format with field labels (e.g. 'm*v^2/2')

    Attributes
    ----------
    experiments: list[dict]
        A list of data for multiple experiments. Each experiment is represented
        as a dictionaty with three elements:
        id: int
            Unique ID of the experiment. Generated automatically in .add_experiment() method
        title: str
            The name of the experiment
        fields: list[dict]
            Fields for collecting and computing data.
            Each field is a dictionary with the following items:
            id: int
                Unique ID of a field. Generated automatically in .add_field() method
            label: str
                Header of a field
            unit: str
                Measure unit of a value, in the actual spreadsheet goes into label,
                Also used in formulas
            type: str
                'gathered' or 'calculated', see above
            error: str | None
                Expression for calculating experimental error of a measured (gathered) value.
                Can be written as excel formula of 'val' (which is the measured value),
                also '%' can be used. 'x%' converts to 'x * 0.01 * val'
            formula: str | None
                Formula for computing a calculated value. Should be written as excel formula
                of other values' labels. e.g. "(m * v^2)/2" if there are fields with labels 'm' and 'v'

    Methods
    -------
    add_experiment(title: str, fields: list = []) -> None
        Add an experiment with provided arguments to *experiments* list

    add_field(experiment, label, unit, field_type, error=None, formula=None) -> None
        Add a field with provided arguments to *fields* list of given experiment

    generate(output_file) -> None
        Returns NotImplementedError if you try to call .generate() method
        from Spreadsheet class (it should be called from one of the generator subclasses)
    """

    # Making an empty list and initializing file type (e.g. 'xlsx' or 'ods') for further file assembling
    # ----------------------------------------
    def __init__(self, filetype: str):

        self.filetype: str = filetype
        self.experiments: list = []
    # ----------------------------------------



    # Add new experiment
    # --------------------------------------------------------------
    def add_experiment(self, title: str = '', fields: list = []):
        
        ID: int = len(self.experiments) + 1  # Auto-incrementing ID

        experiment = {
            'ID': ID,
            'title': title,
            'fields': []
        }

        # If user has provided any fields, we add them
        for field in fields:

            field['ID'] = len(experiment['fields']) + 1  # Auto-incrementing field ID 

            self.__validate_field(field)

            experiment['fields'].append(field)


        self.experiments.append(experiment)
    # --------------------------------------------------------------



    # Add field with specified parameters to specified experiment
    # --------------------------------------------------------------------------------------------
    def add_field(
        self,
        experiment: str | int = 1,  # Either experiment ID or its label.
                                    # By default we select the first experiment
        label: str = '',
        unit: str = '',
        field_type: str = '',
        error=None,
        formula=None
    ) -> None:

        # Extract experiment ID from 'experiment' variable
        exp_ID = self.__get_experiment_ID(experiment)

        # Remember that IDs start with 1, so the position in the list is ID - 1
        exp_pos = exp_ID - 1


        ID: int = len(self.experiments[exp_pos]['fields']) + 1  # Auto-incrementing field ID

        field = {
            'ID': ID,
            'label': label,
            'unit': unit,
            'type': field_type,
            'error': error,
            'formula': formula
        }

        self.__validate_field(field)

        self.experiments[exp_pos]['fields'].append(field)
    # --------------------------------------------------------------------------------------------



    # If someone decides to call .generate() method from Spreadsheet object instead of Generator,
    # we gently remind him of it instead of bashing him in the head with the hammer
    # ------------------------------------------------------------------------------------------
    def generate(self, output_file) -> None:

        del output_file  # Unused, cause we're here to just raise an error

        raise NotImplementedError("This method should be implemented by a generator subclasses "
                                  "(e.g. XLSXGenerator) and not the Spreadsheet class itself")
    # ------------------------------------------------------------------------------------------



    # Getting experiment ID from any given information about it
    # ------------------------------------------------------------------
    def __get_experiment_ID(self, info) -> int:
        
        # If info is an integer, we assume it's an ID
        # and just check if the corresponding experiment exists
        if isinstance(info, int):

            for experiment in self.experiments:

                if experiment['ID'] == info:

                    return experiment['ID']

            raise ValueError(f"No experiment found with ID = {info}")


        # If info is a string, we assume it's a title
        # and do the same procedure
        if isinstance(info, str):

            for experiment in self.experiments:

                if experiment['title'] == info:

                    return experiment['ID']

            raise ValueError(f"No experiment found with title '{info}'")


        # Otherwise, I don't know what you're trying to do
        raise TypeError(f"You're trying to find an experiment "
                        f"providing {info} ({type(info)}).\n"
                        f"Try using experiment ID (int) or label (str)")
    # ------------------------------------------------------------------



    # Validating field items
    # --------------------------------------------------------------------------------
    def __validate_field(self, field) -> None:
        
        # Field must be a dictionary
        if not isinstance(field, dict):
            raise TypeError(f"Field must be a dictionary.\n"
                            f"Got: {type(field)} for field\n"
                            f"{field}") 


        # Field must have an ID
        if 'ID' not in field:
            raise ValueError(f"Field must have an ID. "
                             f"Got field:\n"
                             f"{field}")
        # Field ID must be an integer
        if not isinstance(field['ID'], int):
            raise TypeError(f"Field ID must be an integer.\n"
                            f"Got: {type(field['ID'])} for field\n"
                            f"{field}")


        # Field label must be a non-empty string
        if not isinstance(field['label'], str):
            raise TypeError(f"Field label must be a non-empty string.\n"
                            f"Got: {type(field['label'])} for field\n"
                            f"{field}")

        if not field['label']:
            raise ValueError(f"Field label must not be empty.\n"
                             f"Got: '{field['label']}' for field\n"
                             f"{field}")


        # Field type must be either 'gathered' or 'calculated'
        if not isinstance(field['type'], str):
            raise TypeError(f"Field type must be a non-empty string.\n"
                            f"Got: {type(field['type'])} for field\n"
                            f"{field}")

        if field['type'] not in {"gathered", "calculated"}:
            raise ValueError(f"Field type must be either 'gathered' or 'calculated'.\n"
                             f"Got: '{field['type']}' for field\n"
                             f"{field}")


        # For calculated fields formula must be a non-empty string or number
        if field['type'] == "calculated":

            if not isinstance(field['formula'], (str, int, float)):
                raise TypeError(f"Field formula must be a non-empty string or number.\n"
                                f"Got: {type(field['label'])} for field\n"
                                f"{field}")

            if not field['formula']:
                raise ValueError(f"Field formula must not be empty.\n"
                                 f"Got: '{field['formula']}' for field\n"
                                 f"{field}")

    # --------------------------------------------------------------------------------



    # Nice little string representation in case we ever want to print out our spreadsheet
    # ---------------------------------------------------------------------------------------------
    def __str__(self) -> str:

        string: str = f"{self.__class__}"

        for experiment in self.experiments:

            string += '\n\n'

            header = f"Experiment {experiment['ID']}. {experiment['title']}"

            string += header + '\n'      # Experiment 1. Funny coefficient measurement
            string += '-' * len(header)  # -------------------------------------------

        
            for field in experiment['fields']:

                string += f"\n{field['ID']}. {field['label']}, {field['unit']} ({field['type']}); "
                
                if field['type'] == "gathered":
                    string += f"error: {field['error']}"
                
                elif field['type'] == "calculated":
                    string += f"formula: {field['formula']}"

                # 1. LOL, laughs/s (gathered); error: lsd
                # 2. time, s (gathered); error: 0.1
                # 3. K, laughs (calculated); formula: LOL * time
            

        return string
    # ---------------------------------------------------------------------------------------------

# =====================================================================================================================





"""
Now, as we have the basement, we can make generators which assemble the spreadsheet file.
There are different python libraries for working with different file types,
therefore we need to make an individual generator for each type.
"""
# =====================================================================================================================

class XLSXGenerator(Spreadsheet):
    """
    Spreadsheet subclass.
    Generator which assembles .xlsx file from Spreadsheet experiments list

    Attributes
    ----------
    Same as for Spreadsheet

    Methods
    -------
    generate(output_file: str) -> None
        Assemble .xlsx file in *output_file*.
        *output_file* should have the following formatting:
        '.../Path/to/file/filename'

    See Also
    --------
    Spreadsheet: the basement for creating a spreadsheet.
    """

    def generate(self, output_file) -> None:

        # Initializing the spreadsheet (workbook) and the sheet
        # ----------------------------------------------------------
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # IDK what on Earth must happen for sheet to be None,
        # but pyright really dislikes the lack of this assert
        if sheet is None:
            raise TypeError("The sheet for some reason is None. "
                            "The problem is clearly with openpyxl.")
        # ----------------------------------------------------------



        # Writing experiment titles
        # --------------------------------------------------------------------
        title_row: list = []

        # We want the structure where title cells are merged
        # and experiments are separated by one column e.g.:
        # -------------------------------------------
        # |  Experiment1  |    | Experiment2 |    | ...
        # -------------------------------------------
        # | a | b | c | d |    | e |  f  | g |    | ...
        # -------------------------------------------
        # | 1 | 9 | 0 | 0 |    | 1 | 125 | 8 |    | ...
        # -------------------------------------------
        # | 2 | 5 | 4 | 0 |    | 2 | 900 | 4 |    | ...
        # -------------------------------------------
        #        ...                 ... 

        # To properly merge title cells, we need to keep track of their widths
        column_widths: dict[int, int] = {}  # {column_number: column_width}

        column: int = 1  # In openpyxl column numeration starts with 1


        for experiment in self.experiments:

            title_row.append(f"{experiment['title']}")


            column_width = 0

            for field in experiment['fields']:

                match field['type']:

                    case "gathered":

                        column_width += 2


                    case "calculated":

                        column_width += 1


            # We add n - 1 empty cols to make the total width n,
            # and one more empty col as a separator,
            # which makes the total of n
            title_row.extend([None for _ in range(column_width)])

            column_widths[column] = column_width

            # Keep track of the column number
            column += column_width + 1


        # Now we're left with this
        # ------------------------------------
        # | E1 | | | |    | E2 | | | |    | ...

        # Let's add it to the sheet and then merge needed cells
        sheet.append(title_row)

        for number, width in column_widths.items():
            sheet.merge_cells(f"{get_column_letter(number)}1:"
                              f"{get_column_letter(number + width - 1)}1")

        # --------------------------------------------------------------------



        # Generating and writing headers
        # --------------------------------------------------------------------------------------------------
        header_row: list = []

        # For matching excel column labels to field labels (e.g. m_1 -> A, v -> C etc.)
        # which will be needed to convert formulas into excel format
        column_labels: dict[tuple[int, str], str] = {}  # {(experiment_ID, field_label): column_label}

        column: int = 1  # In openpyxl column numeration starts with 1
        
        
        for experiment in self.experiments:

            for field in experiment['fields']:
                
                # Add the column label (notice that we don't add err field labels 'cause we don't need them)
                column_labels[experiment['ID'], field['label']] = get_column_letter(column)

                match field['type']:

                    case "gathered":

                        # For user input fields we make two columns: value and error
                        header_row.append(f"{field['label']}, {field['unit']}")

                        header_row.append(f"{field['label']} err, {field['unit']}")

                        # Keeping the track of the column number
                        column += 2
            

                    case "calculated":

                        # For calculated fields we make only one column
                        header_row.append(f"{field['label']}, {field['unit']}")

                        # Keeping the track of the column number
                        column += 1


            # Separator
            header_row.append(None)
        

        sheet.append(header_row)  
        # --------------------------------------------------------------------------------------------------



        # Writing rows
        # ---------------------------------------------------------------------------------------------------
        for row_number in range(DATA_BEGINNING_ROW, MAX_ROW_COUNT + DATA_BEGINNING_ROW):
        
            data_row = []
            
            for experiment in self.experiments:

                for field in experiment['fields']:

                    match field['type']:

                        case "gathered":

                            # Leave the column for the gathered data empty
                            data_row.append(None)

                            # Format error formula to a function of 'val'
                            # e.g. 5% + 0.001  ->  0.05 * val + 0.001
                            err_formula = format_error(field['error'])

                            # Replace 'first' with coordinates of the first cell in a field
                            err_formula = err_formula.replace('first',
                                    f"{column_labels[experiment['ID'], field['label']]}{DATA_BEGINNING_ROW}")

                            # Replace 'val' with cell coordinates of gathered data
                            err_formula = err_formula.replace('val',
                                    f"{column_labels[experiment['ID'], field['label']]}{row_number}")

                            # And, finally, write the formula to the cell
                            data_row.append(f"={err_formula}")
                        

                        case "calculated":

                            formula = field["formula"]

                            # Change all labels in the formula to the coordinates of corresponding values
                            for (_, label), col_number in column_labels.items():
                                formula = formula.replace(label, f"{col_number}{row_number}")

                            # Write the formula to the cell
                            data_row.append(f"={formula}")


                # Separator
                data_row.append(None)
                

            sheet.append(data_row)  
        # ---------------------------------------------------------------------------------------------------



        # Saving the spreadsheet
        # ----------------------------------
        workbook.save(output_file + ".xlsx")
        # ----------------------------------





# Error formula formatting. The output is excel-style formula of 'val' and 'first'
# --------------------------------------------------------------------------------------------
def format_error(expr: str) -> str:

    # Formatting percentages
    if "%" in expr:
        expr = expr.replace("%", "*0.01*val")

    # Formatting LSD (last significant digit)
    if "lsd" in expr:
    
        # Arina, please forgive me for what I'm about to do...
        # (TODO: KILL THAT WITH FIRE)
        terrible_nested_statement = ""

        for k in range(-MAX_EXPONENT_ABS, MAX_EXPONENT_ABS + 1):
            terrible_nested_statement += f"IF(ROUND(first, {k}) = first, -{k}, "

        terrible_nested_statement += f"{0}{')' * (MAX_EXPONENT_ABS * 2)}"

        # Okay, so basically this thing is 20 (or MAX_EXPONENT_ABS * 2) nested 'IF' statements
        # each of which checks if the last significant digit has order k.
        # Now let's just raise 10 to the power of k and forget about it.
        expr = expr.replace("lsd", f"IF(val = 0, 0, 10^({terrible_nested_statement}))")

    return expr
# --------------------------------------------------------------------------------------------

# =====================================================================================================================





# Function to select the appropriate generator based on chosen filetype
# (Currently supported: .xlsx)
# ----------------------------------------------------------------------------
def get_spreadsheet_generator(filetype: str) -> XLSXGenerator:

    match filetype:

        case "xlsx":
            return XLSXGenerator(filetype)
    
        # TODO: maybe add .ods support

        case _:
            raise ValueError(f"No generator for file type: {filetype}")
# ----------------------------------------------------------------------------





# Example Usage
# ---------------------------------------------------------------------
filetype = "xlsx"
spreadsheet = get_spreadsheet_generator(filetype)

# Add experiment with fields
spreadsheet.add_experiment(
    title="Ohm's law check",
    fields=[
        {'type': "gathered",   'label': "V",     'unit': "mV",  'error': "3% + 0.01"},
        {'type': "gathered",   'label': "I",     'unit': "mA",  'error': "0.1"},
        {'type': "calculated", 'label': "R_mes", 'unit': "Ohm", 'formula': "V / I"},
        {'type': "calculated", 'label': "R_act", 'unit': "Ohm", 'formula': "20"}
    ]
)

# Add experiment and then fields
spreadsheet.add_experiment(title="Kinetic energy")

spreadsheet.add_field(experiment="Kinetic energy", label="m", unit="kg", field_type="gathered", error="0.04 * lsd")
spreadsheet.add_field(experiment="Kinetic energy", label="v", unit="m/s", field_type="gathered", error="2% + .05")
spreadsheet.add_field(experiment="Kinetic energy", label="K", unit="J", field_type="calculated", formula="m*v^2/2")

print(spreadsheet)

# Setting path to ../output/
path = Path().parent.absolute() / "output"


spreadsheet.generate(f"{path}/spreadsheet")
# ---------------------------------------------------------------------
