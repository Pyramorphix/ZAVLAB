from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment, Side, Border
from copy import copy

# GLOBAL CONSTANTS (TODO: move to config)
# ---------------------------------------------------------------------------------------------------

#  ----------------------
# | Spreadsheet settings |
#  ----------------------

# Maximum row count in the spreadsheet. Feel free to change if needed
MAX_ROW_COUNT = 100

# This is for one very unethical hack in the code for calculating lsd error. (see: format_error())
# Defines a maximum absolute value of the order of numbers in user input fields.
# If equals to x, the program will correctly work with numbers from 10^{-x} to 10^{x+1} - 1.
# I think, no one is crazy enough to type in numbers like 0.0000000005, so 9 is more than enough.
# Feel free to change, but keep in mind that lsd excel formulas' complexity will scale proportionally 
MAX_EXPONENT_ABS = 9

# The row at which data begins in a spreadsheet.
# The first row is for experiment titles, and the second one is for field labels.
# Therefore, we start from the third.
# Do not change unless you clearly understand what you're doing!'
DATA_BEGINNING_ROW = 3



#   -----------
# | Experiments |
#   -----------

# Default amount of rows in experiment
DEFAULT_EXPERIMENT_AMOUNT = 10

# Should we generate experiment's title automatically if no title is specified,
# or just leave it blank
AUTO_EXPERIMENT_TITLES = True



#   -----
# | Fonts |
#   -----

# Font size of experiment titles (pretty self-explanatory)
TITLE_FONT_SIZE = 28

# Font size of everything else
DEFAULT_FONT_SIZE = 22

# Bold font settings
IS_TITLE_FONT_BOLD = True
IS_LABEL_FONT_BOLD = True
IS_DATA_FONT_BOLD = False

# Font family
TITLE_FONT = "Calibri"
DEFAULT_FONT = "Calibri"



#   --------
# | Coloring |
#   --------

# Cell background colors
TITLE_BG_COLOR = "dddddd"
CONST_BG_COLOR = "afd095"
GATHERED_BG_COLOR = "ec9ba4"
CALCULATED_BG_COLOR = "ffffa6"

# Cell font colors
TITLE_FONT_COLOR = "000000"
CONST_FONT_COLOR = "000000"
GATHERED_FONT_COLOR = "000000"
CALCULATED_FONT_COLOR = "000000"

# ---------------------------------------------------------------------------------------------------





"""
This part was made by Vlad.
In this file, all comments in Russian must be treated as funny jokes!
Comments in English are intended to be helpful (but can be also treated as jokes if you want).
If you'll find comments in any other language, immediately report to the Waste Management!
(Sentinel Prime said he will reward you for this)

Now, let's go!
"""





"""
Here we define all the styles (cell colors, font sizes etc.)
"""
# =====================================================================================================================

# Border presets
thick_bd = Side(style="thick", color="000000")
bd = Side(style="medium", color="000000")
thin_bd = Side(style="thin", color="000000")

# Experiment title
# --------------------------------------------------------------
title_style = NamedStyle(
    name="title",
    font=Font(name=TITLE_FONT,
              size=TITLE_FONT_SIZE,
              color=TITLE_FONT_COLOR,
              bold=IS_TITLE_FONT_BOLD),

    fill=PatternFill(start_color=TITLE_BG_COLOR,
                     end_color=TITLE_BG_COLOR,
                     fill_type="solid"),

    border=Border(
        left=thick_bd,
        top=thick_bd,
        right=thick_bd,
        bottom=thick_bd,
    ),

    alignment=Alignment(horizontal="center", vertical="center"),
)
# --------------------------------------------------------------


# Data field labels
# ------------------------------------------------------------------------
label_style = NamedStyle(
    name="label",
    font=Font(name=DEFAULT_FONT,
              size=DEFAULT_FONT_SIZE,
              bold=IS_LABEL_FONT_BOLD),

    border=Border(
        left=bd,
        top=bd,
        right=bd,
        bottom=bd,
    ),

    alignment=Alignment(horizontal="center", vertical="center"),
)

const_label_style = copy(label_style)
const_label_style.name = "const_label"
const_label_style.font.color = CONST_FONT_COLOR
const_label_style.alignment.horizontal = "right"
const_label_style.border.bottom = None
const_label_style.fill = PatternFill(start_color=CONST_BG_COLOR,
                                     end_color=CONST_BG_COLOR,
                                     fill_type="solid")

gathered_label_style = copy(label_style)
gathered_label_style.name = "gathered_label"
gathered_label_style.font.color = GATHERED_FONT_COLOR
gathered_label_style.fill = PatternFill(start_color=GATHERED_BG_COLOR,
                                        end_color=GATHERED_BG_COLOR,
                                        fill_type="solid")

calculated_label_style = copy(label_style)
calculated_label_style.name = "calculated_label"
calculated_label_style.font.color = CALCULATED_FONT_COLOR
calculated_label_style.fill = PatternFill(start_color=CALCULATED_BG_COLOR,
                                          end_color=CALCULATED_BG_COLOR,
                                          fill_type="solid")
# ------------------------------------------------------------------------


# Data field contents
# -----------------------------------------------------------------------
data_style = NamedStyle(
    name="data",
    font=Font(name=DEFAULT_FONT,
              size=DEFAULT_FONT_SIZE,
              bold=IS_DATA_FONT_BOLD),

    border=Border(
        left=thin_bd,
        right=thin_bd,
        bottom=thin_bd,
    ),
    
    alignment=Alignment(horizontal="right", vertical="center"),
)

const_data_style = copy(data_style)
const_data_style.name = "const_data"
const_data_style.font.color = CONST_FONT_COLOR
const_data_style.border.left = bd
const_data_style.border.right = bd
const_data_style.border.bottom = bd
const_data_style.fill = PatternFill(start_color=CONST_BG_COLOR,
                                     end_color=CONST_BG_COLOR,
                                     fill_type="solid")

gathered_data_style = copy(data_style)
gathered_data_style.name = "gathered_data"
gathered_data_style.font.color = GATHERED_FONT_COLOR
gathered_data_style.fill = PatternFill(start_color=GATHERED_BG_COLOR,
                                        end_color=GATHERED_BG_COLOR,
                                        fill_type="solid")

calculated_data_style = copy(data_style)
calculated_data_style.name = "calculated_data"
calculated_data_style.font.color = CALCULATED_FONT_COLOR
calculated_data_style.fill = PatternFill(start_color=CALCULATED_BG_COLOR,
                                          end_color=CALCULATED_BG_COLOR,
                                          fill_type="solid")

# -----------------------------------------------------------------------

# =====================================================================================================================



"""
Here we make the basement for out spreadsheet.
"""
# =====================================================================================================================
class Field:
    """
    A column in a spreadsheet.

    Each field is either for user input (type='gathered') or for calculated values (type='calculated').
    For the convenience, there are also constant fields (type='const'),
    which can be pre-defined in the code, or typed in the spreadsheet by the user.
    They are displayed in the first column of every experiment and have only one cell each.
    Every field has the label and measure unit.

    User input fields additionally have the 'error' element for entering the experimental error.
    It can be either a certain value (e.g. '0.005') or a formula depending on measured value (e.g. '2% + 0.04')

    Calculated values have the 'formula' element for the calculation formula.
    It should be typed in using excel-readable format with field labels (e.g. 'm*v^2/2')

    Constant fields have the 'value' element for the constant value.
    By default it's None, which leads to an empty cell.


    Attributes
    ----------
    id: int
        Unique ID of a field. Generated automatically in .add_field() method

    label: str
        Header of a field

    unit: str
        Measure unit of a value, in the actual spreadsheet goes into label,
        Also used in formulas

    field_type: str
        "gathered" or "calculated", see above

    error: int | float | str | None
        Expression for calculating experimental error of a measured (gathered) value.
        Can be written as excel formula of 'val' (which is the measured value),
        '%' can be used. 'x%' converts to 'x * 0.01 * val'.
        'lsd' can be used. 'lsd' converts to least significant digit of the first (!!) element of the field
        Used only for "const" or "gathered" fields

    formula: int | float | str | None
        Formula for computing a calculated value. Should be written as excel formula
        of other values' labels. e.g. "(m * v^2)/2" if there are fields with labels 'm' and 'v'.
        Used only for "calculated" fields.

    value: int | float | str | None
        The value of the constant field.
        Used only for "const" fields.
    """

    # --------------------------------------------------------------------
    def __init__(
        self,
        label: str = '',
        field_type: str = '',
        unit: str = "",
        error: int | float | str | None = None,
        formula: int | float | str | None = None,
        value: int | float | str | None = None,
        id: int = 0,
    ) -> None:

        self.label: str = label
        self.field_type: str = field_type
        self.unit: str = unit
        self.error: int | float | str | None = error
        self.formula: int | float | str | None = formula
        self.value: int | float | str | None = value
        self.id: int = id

        # Set int const error to least significant digit by default
        if field_type == "const" and isinstance(value, int) and not error:
            self.error = "lsd"
    # --------------------------------------------------------------------



    # ------------------------------------------------------------------------------------------------
    def validate(self) -> None:
        """
        Field validation

        Also called by Experiment.validate()

        Raises
        ------
        TypeError with verbose information
            if any parameter of the field has wrong type.

        ValueError with verbose information
            if any parameter of the field has wrong value.
        """
        
        # Field ID must be an integer > 0
        if not isinstance(self.id, int):
            raise TypeError(f"Field ID must be an integer.\n"
                            f"Got: {type(self.id)} for field\n"
                            f"{self}")

        if self.id <= 0:
            raise ValueError(f"Field ID must be > 0.\n"
                             f"Got: {self.id} for field\n"
                             f"{self}")


        # Field label must be a non-empty string
        if not isinstance(self.label, str):
            raise TypeError(f"Field label must be a non-empty string.\n"
                            f"Got: {type(self.label)} for field\n"
                            f"{self}")

        if not self.label:
            raise ValueError(f"Field label must not be empty.\n"
                             f"Got: '{self.label}' for field\n"
                             f"{self}")


        # Field type must be either "gathered", "calculated" or "const"
        if self.field_type not in {"gathered", "calculated", "const"}:
            raise ValueError(f"Field type must be either \"gathered\", \"calculated\" or \"const\".\n"
                             f"Got: '{self.field_type}' for field\n"
                             f"{self}")


        # For calculated fields formula must be a non-empty string or number
        if self.field_type == "calculated":

            if not isinstance(self.formula, (str, int, float)):
                raise TypeError(f"Field formula must be a non-empty string or number.\n"
                                f"Got: {type(self.formula)} for field\n"
                                f"{self}")

            if not self.formula:
                raise ValueError(f"For calculated field, formula must not be empty.\n"
                                 f"Got: '{self.formula}' for field\n"
                                 f"{self}")
    # ------------------------------------------------------------------------------------------------



    # Nice string representation
    # ------------------------------------------------------------------------------
    def __str__(self) -> str:

        string: str = f"ZAVLAB {self.__class__.__name__}"
        
        if self.unit:
            string += f"\n{self.id}. {self.label}, {self.unit} ({self.field_type}); "
        else:
            string += f"\n{self.id}. {self.label} ({self.field_type}); "

                
        match self.field_type:
            case "gathered":
                string += f"error: {self.error}"
        
            case "calculated":
                string += f"formula: {self.formula}"

            case "constant":
                string += f"value: {self.value}, error: {self.error}"

            case _:
                pass


        return string
    # ------------------------------------------------------------------------------





class Experiment:
    """
    A collection of fields for one experiment.

    Attributes
    ----------
    id: int
        Unique ID of the experiment. Generated automatically in .add_experiment() method

    title: str
        The name of the experiment.
        Default is "Experiment {id}"

    fields: list[Field]
        Fields for collecting and computing data.

    constants: list[Field]
        Fields with type="const". They're just much easier to process when in a separate list.

    amount: int
        Amount of experiment repititions (rows in the spreadsheet).
        Default is DEFAULT_EXPERIMENT_AMOUNT
    """

    # --------------------------------------------------------------------------
    def __init__(
        self,
        title: str = "",
        fields: list[Field] | None = None,
        constants: list[Field] | None = None,
        amount: int = DEFAULT_EXPERIMENT_AMOUNT,
        id: int = 0,
    ) -> None:
        
        self.title: str = title
        self.fields: list[Field] = fields if fields is not None else []
        self.constants: list[Field] = constants if constants is not None else []
        self.amount = amount
        self.id: int = id

        # Set experiment title to "Experiment {ID}" by default:
        if not title and AUTO_EXPERIMENT_TITLES:
            self.title = f"Experiment {id}"
    # --------------------------------------------------------------------------



    # --------------------------------------------------------------------------------
    def add_field(
        self,
        field: Field | None = None,
        label: str = '',
        unit: str = '',
        field_type: str = '',
        error: int | float | str | None = None,
        formula: int | float | str | None = None,
        value: int | float | str | None = None,
    ) -> None:
        """
        Add field with specified parameters to the experiment.

        Parameters
        ----------
        Either Field object,
        or all the attributes of the FIeld class except ID (generated automatically):
        label, unit (opt), field_type, error, formula, value

        Notes
        -----
        Calls Field.validate() for data validation
        """

        # If we don't receive a Field object,
        # assemble it from given parameters
        if not field:
            field = Field(
                label=label,
                unit=unit,
                field_type=field_type,
                error=error,
                formula=formula,
                value=value,
            )

        self.__generate_field_id(field)

        field.validate()

        self.__append_field(field)
    # --------------------------------------------------------------------------------



    # ----------------------------------------------------------------------
    def validate(self):
        """
        Experiment validation.

        Raises
        ------
        TypeError with verbose information
            if any parameter of the experiment or its field has wrong type.

        ValueError with verbose information
            if any parameter of the experiment or its field has wrong value.

        Notes
        -----
        Calls Field.validate() for field validation
        """

        # Experiment ID must be an integer > 0
        if not isinstance(self.id, int):
            raise TypeError(f"Experiment ID must be an integer.\n"
                            f"Got: {type(self.id)} for experiment\n"
                            f"{self}")

        if self.id <= 0:
            raise ValueError(f"Experiment ID must be > 0.\n"
                             f"Got: {self.id} for experiment\n"
                             f"{self}")


        # Experiment amount must be an integer > 0
        if not isinstance(self.amount, int):
            raise TypeError(f"Experiment amount must be an integer.\n"
                            f"Got: {type(self.amount)} for experiment\n"
                            f"{self}")

        if self.amount <= 0:
            raise ValueError(f"Experiment amount must be > 0.\n"
                             f"Got: {self.amount} for experiment\n"
                             f"{self}")


        # Validate fields
        for field in self.fields:
            field.validate()

        for const in self.constants:
            const.validate()

    # ----------------------------------------------------------------------


    
    # Auto-incrementing field ID (constants have their own ID system)
    # ------------------------------------------
    def __generate_field_id(self, field: Field):

        if field.field_type == "const":
            field.id = len(self.constants) + 1  

        else:
            field.id = len(self.fields) + 1  
    # ------------------------------------------



    # Append field to either fields or constants based on its type
    # -------------------------------------
    def __append_field(self, field: Field):

        if field.field_type == "const":
            self.constants.append(field)

        else:
            self.fields.append(field)
    # -------------------------------------



    # Nice string representation
    # ----------------------------------------------------------------------------
    def __str__(self) -> str:
        
        string: str = f"ZAVLAB {self.__class__.__name__}\n"

        header = f"Experiment {self.id}. {self.title}"

        string += header + '\n'      # Experiment 1. Funny coefficient measurement
        string += '-' * len(header)  # -------------------------------------------


        if not self.constants:
            string += "\nNo constants.\n"  # НЕ БРАТЬ КОНСТА!

        else:
            string += "\nConstants:\n"

        for const in self.constants:

            const_repr = const.__str__()

            # Remove the first line ("ZAVLAB Field") from const_repr 
            string += const_repr.split('\n', 1)[1]
            string += '\n'

            # Constants
            # 1. Konst (const); value: NO  (НЕ БРАТЬ КОНСТА!)

        
        if not self.fields:
            string += "\nNo fields.\n"

        else:
            string += "\nFields:\n"

        for field in self.fields:

            field_repr = field.__str__()

            # Remove the first line ("ZAVLAB Field") from field_repr 
            string += field_repr.split('\n', 1)[1]
            string += '\n'

            # Fields
            # 1. LOL, laughs/s (gathered); error: lsd
            # 2. time, s (gathered); error: 0.1
            # 3. K, laughs (calculated); formula: LOL * time

        return string
    # ----------------------------------------------------------------------------





class Spreadsheet:
    """
    The basement for creating a spreadsheet. Basically a list of experiments.

    Attributes
    ----------
    experiments: list[Experiment]
        A list of data for multiple experiments.        

    Methods
    -------
    add_experiment(Experiment object or
                   title, fields=[], constants=[], amount) -> None
        Add an experiment with provided arguments to *experiments* list

    add_field(experiment,
              Field object or
              label, unit, field_type, error=None, formula=None, value=None) -> None
        Add a field with provided arguments to *fields* list of given experiment.
        experiment variable is either title or ID 

    generate(output_file) -> None
        Returns NotImplementedError if you try to call .generate() method
        from Spreadsheet class (it should be called from one of the generator subclasses)
    """

    # -------------------------------------------------------------------------------------
    def __init__(self, experiments: list[Experiment] | None = None):

        self.experiments: list[Experiment] = experiments if experiments is not None else []
    # -------------------------------------------------------------------------------------



    # --------------------------------------------------------------
    def add_experiment(
        self,
        experiment: Experiment | None = None,
        title: str = '',
        fields: list[Field] | None = None,
        constants: list[Field] | None = None,
        amount: int = DEFAULT_EXPERIMENT_AMOUNT,
    ) -> None:
        """
        Add new experiment with specified parameters to the spreadsheet.

        Parameters
        ----------
        Either the Experiment object,
        or all the attributes of Experiment class except ID (generated automatically):
        title, fields (opt), constants (opt), amount

        Notes
        -----
        Calls Experiment.validate() for data validation
        """

        # If we don't receive an Experiment object,
        # assemble it from given parameters
        if experiment is None:
            experiment = Experiment(
                title=title,
                fields=fields if fields is not None else [],
                constants=constants if constants is not None else [],
                amount=amount,
            )

        self.__generate_experiment_id(experiment)

        experiment.validate()

        self.__append_experiment(experiment)
    # --------------------------------------------------------------



    # Add field with specified parameters to specified experiment
    # --------------------------------------------------------------------------------------------
    def add_field(
        self,
        experiment: str | int,  # Either experiment ID or its label.
        field: Field | None = None,
        label: str = '',
        unit: str = '',
        field_type: str = '',
        error: int | float | str | None = None,
        formula: int | float | str | None = None,
        value: int | float | str | None = None,
    ) -> None:
        """
        Add new field with specified parameters to specified experiment

        Parameters
        ----------
        experiment: str | int
            Either experiment ID or its label
        field (Field object) or all the attributes of Field class except ID (generated automatically):
        label, unit (opt), field_type, error (opt), formula (opt), value (opt)

        Notes
        -----
        Calls Experiment.add_field to do the job
        """

        # Extract experiment ID from 'experiment' variable
        exp_ID = self.__get_experiment_id(experiment)

        # Remember that IDs start with 1, so the position in the list is ID - 1
        exp_pos = exp_ID - 1

        self.experiments[exp_pos].add_field(field, label, unit, field_type, error, formula, value)
    # --------------------------------------------------------------------------------------------



    # If someone decides to call .generate() method from Spreadsheet object instead of Generator,
    # we gently remind him of it instead of bashing him in the head with the hammer
    # ------------------------------------------------------------------------------------------
    def generate(self, output_file: str) -> None:
        """
        This method must be implemented by generator subclasses
        (e.g. XLSXGenerator) and not the Spreadsheet class itself!
        """

        del output_file  # Unused, cause we're here to just raise an error

        raise NotImplementedError("This method should be implemented by generator subclasses "
                                  "(e.g. XLSXGenerator) and not the Spreadsheet class itself")
    # ------------------------------------------------------------------------------------------



    # Getting experiment id from any given information about it
    # ------------------------------------------------------------------
    def __get_experiment_id(self, info: int | str) -> int:

        # If info is an integer, we assume it's an ID
        # and just check if the corresponding experiment exists
        if isinstance(info, int):

            for experiment in self.experiments:
                if experiment.id == info:
                    return experiment.id

            raise ValueError(f"No experiment found with ID = {info}")


        # If info is a string, we assume it's a title
        # and do the same procedure
        if isinstance(info, str):

            for experiment in self.experiments:
                if experiment.title == info:
                    return experiment.id

            raise ValueError(f"No experiment found with title '{info}'")


        # Otherwise, I don't know what you're trying to do
        raise TypeError(f"You're trying to find an experiment "
                        f"providing: {info} ({type(info)}).\n"
                        f"ZAVLAB doesn't understand this. "
                        f"Try using experiment ID (int) or label (str)")
    # ------------------------------------------------------------------


    # Auto-incrementing experiment ID 
    # ---------------------------------------------------------
    def __generate_experiment_id(self, experiment: Experiment):

            experiment.id = len(self.experiments) + 1  
    # ---------------------------------------------------------



    # Append experiment to experiments 
    # ----------------------------------------------------
    def __append_experiment(self, experiment: Experiment):

            self.experiments.append(experiment)
    # ----------------------------------------------------



    # Nice little string representation in case we ever want to print out our spreadsheet
    # ---------------------------------------------------------------------------------------------
    def __str__(self) -> str:

        string: str = f"ZAVLAB {self.__class__.__name__}"

        if not self.experiments:
            string += "\nNo experiments"

        else:
            for experiment in self.experiments:

                string += "\n\n"

                experiment_repr = experiment.__str__()

                # Remove the first line ("ZAVLAB Experiment") from experiment_repr 
                string += experiment_repr.split('\n', 1)[1]

        return string
    # ---------------------------------------------------------------------------------------------

# =====================================================================================================================





"""
Now, as we have the basement, we can make generators which assemble the spreadsheet file.
There are different python libraries for working with different file types,
therefore we need to make an individual generator for each type.
"""
# =====================================================================================================================

class XLSXGenerator(Spreadsheet):
    """
    Spreadsheet subclass.
    Generator which assembles .xlsx file from Spreadsheet experiments list

    Attributes
    ----------
    Same as for Spreadsheet

    Methods
    -------
    generate(output_file: str) -> None
        Assemble .xlsx file in *output_file*.
        *output_file* should have the following formatting:
        '.../Path/to/file/filename'

    See Also
    --------
    Spreadsheet: the basement for creating a spreadsheet.
    """

    # --------------------------------------------------------------
    def __init__(self, experiments: list[Experiment] | None = None):

        super().__init__(experiments)

        # Initialize the spreadsheet (workbook) and the sheet
        workbook = Workbook()
        sheet = workbook.active

        # IDK what on Earth must happen for sheet to be None,
        # but pyright really dislikes the lack of this assert
        if sheet is None:
            raise TypeError("The sheet for some reason is None. "
                            "The problem is clearly with openpyxl.")

        self.workbook = workbook
        self.sheet = sheet
    # --------------------------------------------------------------



    # ---------------------------------------------------------------------
    def generate(self, output_file: str) -> None:
        """
        Assemble .xlsx spreadsheet from experiments list in output_file
    
        Parameters
        ----------
        output_file: str
            File path to desired output file e.g. "Path/to/file/spreadsheet"
            Must not contain extension (.xlsx)
            Passed to openpyxl.Workbook.save()

        Notes
        -----
        Experiments are generated under each other with one row between them
        """

        # We want the structure where title cells are merged
        # and experiments are separated by one column e.g.:
        # ----------------------------------------------------------------
        # |            Experiment1            |    | Experiment2 |    | ...
        # ----------------------------------------------------------------
        # | const | a, smth | b, m | b err, m |    | c |  d  | e |    | ...
        # ----------------------------------------------------------------
        # | l, cm |    2    | 1984 |    20    |    | 1 | 125 | 3 |    | ...
        # ----------------------------------------------------------------
        # | 80085 |    3    | 2000 |    21    |    | 2 | 900 | 3 |    | ...
        # ----------------------------------------------------------------
        #     ^           ...                        ^   ...
        #     |                                      |
        # Constants                             No constants
        #                                     (НЕ БРАТЬ КОНСТА!)



        # Register styles (in openpyxl we need to do it before using them)
        styles = [
            title_style,
            label_style,
            const_label_style,
            gathered_label_style,
            calculated_label_style,
            data_style,
            const_data_style,
            gathered_data_style,
            calculated_data_style,
        ]

        for style in styles:
            self.workbook.add_named_style(style)


        # Write experiments under each other
        # TODO: Add more flexibility to this
        row = 1
        col = 1

        for experiment in self.experiments:
            self.__wirte_experiment(experiment, row, col)

            # Headers
            row += 2

            # Values
            row += max(experiment.amount, len(experiment.constants) * 2)

            # Separator
            row += 1 


        # Save the spreadsheet
        self.workbook.save(output_file + ".xlsx")
    # ---------------------------------------------------------------------



    # -----------------------------------------------------------------------------------
    def __wirte_experiment(self, experiment: Experiment, start_row: int, start_col: int):
    
        # Keep track of row and col in the process
        row: int = start_row
        col: int = start_col

        # Write title
        self.sheet.cell(row, col, experiment.title).style = "title"
        row += 1

        # We need these to replace field labels in formulas with cell coordinates
        field_label_to_col: dict[str, int] = {}
        const_label_to_cell: dict[str, str] = {}


        # Write constants
        if experiment.constants:
            self.__write_constant_sector(experiment.constants, row, col)

            row += 2
            for const in experiment.constants:
                const_label_to_cell[const.label] = f"{get_column_letter(col)}{row}"
                row += 2
            
            # Reset row counter
            row -= 2 * (len(experiment.constants) + 1)

            col += 2


        # For resetting col counter in the future
        col_bak = col

        # Write Fields
        if experiment.fields:
            # First, we need to cycle through all fields to assign columns to labels
            for field in experiment.fields:
                field_label_to_col[field.label] = col
                col += 1
                if field.field_type == "gathered" and field.error is not 0:
                    col += 1

            # Reset col counter
            col = col_bak

            for field in experiment.fields:
                self.__write_field(field, row, col, experiment.amount,
                                   field_label_to_col, const_label_to_cell)
                col += 1
                # Gathered field with error takes 2 columns
                if field.field_type == "gathered" and field.error is not 0:
                    col += 1
            col -= 1

        # Merge title cell
        self.sheet.merge_cells(start_row=start_row, start_column=start_col,
                               end_row=start_row, end_column=col)
    # -----------------------------------------------------------------------------------



    # ----------------------------------------------------------------------------------
    def __write_constant_sector(self, constants: list[Field], start_row: int, col: int):

        # Keep track of row in the process
        row: int = start_row

        # Write header
        self.sheet.cell(row, col, "Constants").style = "const_label"
        self.sheet.cell(row, col + 1, "err").style = "const_label"
        self.sheet.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=col + 1)
        row += 1

        for const in constants:

            # Write label
            self.sheet.cell(row, col, format_label(const)).style = "const_label"
            # Write error label
            self.sheet.cell(row, col + 1, "err").style = "const_label"
            row += 1

            # Write value
            self.sheet.cell(row, col, const.value).style = "const_data"
            # Write error
            err = format_error(const.error)
            if err is not None:
                err = err.replace("first", "val")
                err = err.replace("val", f"{get_column_letter(col)}{row}")
                err = '=' + err
            self.sheet.cell(row, col + 1, err).style = "const_data"
            row += 1
    # ----------------------------------------------------------------------------------



    # ------------------------------------------------------------------------------------------------------
    def __write_field(
            self,
            field: Field,
            start_row: int,
            col: int,
            amount: int,
            field_label_to_col,
            const_label_to_cell
    ) -> None:

        # Keep track of row in process
        row: int = start_row

        # Write label
        self.sheet.cell(row, col, format_label(field)).style = f"{field.field_type}_label"
        if field.field_type == "gathered" and field.error is not 0:
            # Write err label
            self.sheet.cell(row, col + 1, format_label(field, err=True)).style = f"{field.field_type}_label"

        row += 1


        if field.field_type == "gathered":

            # Write empty cells
            for _ in range(amount):
                self.sheet.cell(row, col, None).style = "gathered_data"
                row += 1

            # Reset row counter
            row -= amount

            # Write errors
            if field.error is not 0:
                for _ in range(amount):
                    err = format_error(field.error)
                    if err is not None:
                        err = err.replace("first", f"{get_column_letter(col)}{start_row + 1}")
                        err = err.replace("val", f"{get_column_letter(col)}{row}")
                        err = '=' + err
                    self.sheet.cell(row, col + 1, err).style = "gathered_data"
                    row += 1


        elif field.field_type == "calculated":

            # Write formulas
            for _ in range(amount):
                formula = format_formula(field.formula, row, field_label_to_col, const_label_to_cell)
                self.sheet.cell(row, col, f"={formula}").style="calculated_data"
                row += 1
    # ------------------------------------------------------------------------------------------------------








# Error formula formatting. The output is excel-style formula of 'val' and 'first'
# --------------------------------------------------------------------------------------------
def format_error(error: int | float | str | None) -> str | None:

    if error is None:
        return None

    expr = str(error)
    # Formatting percentages
    if "%" in expr:
        expr = expr.replace("%", "*0.01*val")

    # Formatting LSD (last significant digit)
    if "lsd" in expr:

        # Arina, please forgive me for what I'm about to do...
        # (TODO: KILL THAT WITH FIRE)
        terrible_nested_statement = ""

        for k in range(-MAX_EXPONENT_ABS, MAX_EXPONENT_ABS + 1):
            terrible_nested_statement += f"IF(ROUND(first, {k}) = first, -{k}, "

        terrible_nested_statement += f"{0}{')' * (MAX_EXPONENT_ABS * 2)}"

        # Okay, so basically this thing is 20 (or MAX_EXPONENT_ABS * 2) nested 'IF' statements
        # each of which checks if the last significant digit has order k.
        # Now let's just raise 10 to the power of k and forget about it.
        expr = expr.replace("lsd", f"IF(val = 0, 0, 10^({terrible_nested_statement}))")

    return expr
# --------------------------------------------------------------------------------------------



# Replace all field labels in formula with cell coordinates
# -------------------------------------------------------------------------------------------------------
def format_formula(formula: int | float | str | None, row: int, field_label_to_col, const_label_to_cell):

    formatted_formula = str(formula)

    for label, col in field_label_to_col.items():
        formatted_formula = formatted_formula.replace(label, f"{get_column_letter(col)}{row}")

    for label, cell in const_label_to_cell.items():
        formatted_formula = formatted_formula.replace(label, cell)

    return formatted_formula
# -------------------------------------------------------------------------------------------------------



# Nice label formatting
# -------------------------------------------------------
def format_label(field: Field, err: bool = False) -> str:
    
    label = field.label
    unit = field.unit

    formatted_label = f"{label}"

    if err:
        formatted_label += " err"

    if unit:
        formatted_label += f", {unit}"

    return formatted_label
# -------------------------------------------------------

# =====================================================================================================================





# Function to select the appropriate generator based on chosen filetype
# (Currently supported: .xlsx)
# ----------------------------------------------------------------------------
def get_spreadsheet_generator(filetype: str = "xlsx") -> XLSXGenerator:
    """
    Select the appropriate generator object based on chosen file type

    Parameters
    ----------
    filetype: str = "xlsx"
        Generated spreadsheet file type.
        Currently supported: xlsx

    Returns
    -------
    Generator object based on the file type:
    "xlsx" -> XLSXGenerator

    Raises
    ------
    ValueError
        If given file type is not supported
    """

    match filetype:

        case "xlsx":
            return XLSXGenerator()

        # TODO: maybe add .ods support

        case _:
            raise ValueError(f"No generator for file type: {filetype}")
# ----------------------------------------------------------------------------

